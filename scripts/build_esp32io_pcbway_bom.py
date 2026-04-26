#!/usr/bin/env python3
"""
Build a PCBWay-style BOM xlsx for esp32io-v1.01 by merging:
  - KiCad Interactive BOM embedded in pcb-BOM.html (from the Gerber zip)
  - Mouser saved cart .xls (Mouser #, Mfr #, Manufacturer, Description)

Gerber copper/mask files do not contain BOM data; this script uses pcb-BOM.html
from the same manufacturing archive when available.
"""

from __future__ import annotations

import argparse
import json
import re
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

HEADER = (
    "Item #",
    "*Designator",
    "*Qty",
    "Manufacturer",
    "*Mfg Part #",
    "Description / Value",
    "*Package/Footprint ",
    "Type",
    "Your Instructions / Notes",
)


def _ensure_deps() -> None:
    import os

    try:
        import lzstring  # noqa: F401
        import openpyxl  # noqa: F401
        import xlrd  # noqa: F401
    except ImportError:
        pass
    else:
        return
    venv = Path(tempfile.mkdtemp(prefix="bom_build_venv_"))
    subprocess.run([sys.executable, "-m", "venv", str(venv)], check=True)
    pip = venv / "bin" / "pip"
    subprocess.run([str(pip), "install", "-q", "lzstring", "openpyxl", "xlrd"], check=True)
    os_exec = venv / "bin" / "python3"
    if not os_exec.exists():
        os_exec = venv / "bin" / "python"
    argv = [str(os_exec), str(Path(__file__).resolve()), *sys.argv[1:]]
    env = {**os.environ, "BOM_BUILD_VENV_CLEANUP": str(venv)}
    os.execve(str(os_exec), argv, env)


def _maybe_cleanup_venv() -> None:
    import os

    p = os.environ.get("BOM_BUILD_VENV_CLEANUP")
    if p and Path(p).exists():
        shutil.rmtree(p, ignore_errors=True)


def load_cart(path: Path) -> dict[str, dict[str, str]]:
    import xlrd

    wb = xlrd.open_workbook(str(path))
    sh = wb.sheet_by_index(0)
    # Row 6 in file is header (0-based row 6): Mouser #, Mfr. #, Manufacturer, ...
    header_row = None
    for r in range(sh.nrows):
        row = [str(sh.cell_value(r, c)).strip() for c in range(min(4, sh.ncols))]
        if row and row[0] == "Mouser #":
            header_row = r
            break
    if header_row is None:
        raise SystemExit(f"Could not find Mouser header row in {path}")
    out: dict[str, dict[str, str]] = {}
    for r in range(header_row + 1, sh.nrows):
        mouser = str(sh.cell_value(r, 0)).strip()
        if not mouser:
            continue
        mfr_part = str(sh.cell_value(r, 1)).strip()
        mfr_name = str(sh.cell_value(r, 2)).strip()
        desc = str(sh.cell_value(r, 4)).strip()
        out[mouser] = {
            "mfr_part": mfr_part,
            "manufacturer": mfr_name,
            "description": desc,
        }
    return out


def load_ibom_rows(html_path: Path) -> list[tuple[int, str, str, list[str], list[str]]]:
    import lzstring

    html = html_path.read_text(encoding="utf-8", errors="replace")
    m = re.search(r'LZString\.decompressFromBase64\("([^"]+)"\)', html)
    if not m:
        raise SystemExit(f"No iBOM payload in {html_path}")
    raw = lzstring.LZString().decompressFromBase64(m.group(1))
    if raw is None:
        raise SystemExit("LZString decompress failed")
    data = json.loads(raw)
    rows_out: list[tuple[int, str, str, list[str], list[str]]] = []
    for row in data["bom"]["both"]:
        qty, value, fp, refs_block, extra = row
        refs = sorted([r[0] for r in refs_block], key=_ref_sort_key)
        extras = [str(x).strip() for x in extra if str(x).strip()]
        rows_out.append((int(qty), str(value), str(fp), refs, extras))
    return rows_out


def _ref_sort_key(ref: str) -> tuple[int, str]:
    m = re.match(r"^([A-Za-z]+)(\d+)$", ref)
    if m:
        return (0, f"{m.group(1)}{int(m.group(2)):06d}")
    return (1, ref)


def guess_footprint_type(footprint: str) -> str:
    fp = footprint.upper()
    th_tokens = (
        "PINHEADER",
        "PINSOCKET",
        "RJ45",
        "TERMINALBLOCK",
        "TTGO",
        "MODULE",
    )
    if any(t in fp.replace("_", "") for t in th_tokens):
        return "Through Hole"
    return "SMD"


def resolve_cart_keys(extras: list[str], cart: dict[str, dict[str, str]]) -> dict[str, str] | None:
    if not extras:
        return None
    blob = extras[0]
    candidates = [c.strip() for c in re.split(r"\s+or\s+", blob, flags=re.I) if c.strip()]
    for c in candidates:
        if c in cart:
            return cart[c]
    # Longest Mouser key contained in blob
    for key in sorted(cart.keys(), key=len, reverse=True):
        if key in blob:
            return cart[key]
    return None


def write_pcbway_xlsx(
    out_path: Path,
    title: str,
    lines: list[tuple[int, int, str, str, str, str, str, str, str]],
) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["D2"] = title
    for col, name in enumerate(HEADER, start=1):
        ws.cell(row=6, column=col, value=name)
    for row in lines:
        item, qty, designators, manufacturer, mpn, desc, fp, typ, notes = row
        ws.append(
            [
                item,
                designators,
                qty,
                manufacturer,
                mpn,
                desc,
                fp,
                typ,
                notes,
            ]
        )
    wb.save(out_path)


def main() -> None:
    if "BOM_BUILD_VENV_CLEANUP" not in __import__("os").environ:
        _ensure_deps()

    import lzstring  # noqa: F401
    import openpyxl  # noqa: F401
    import xlrd  # noqa: F401

    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--zip",
        type=Path,
        default=Path(__file__).resolve().parents[1]
        / "esp32io-v1.01_PCBWay Community.zip",
    )
    ap.add_argument(
        "--cart",
        type=Path,
        default=Path(__file__).resolve().parents[1] / "SavedCart_Apr26_1200PM.xls",
    )
    ap.add_argument(
        "--out",
        type=Path,
        default=Path(__file__).resolve().parents[1] / "esp32io-v1.01_BOM_PCBWay.xlsx",
    )
    args = ap.parse_args()

    with tempfile.TemporaryDirectory() as td:
        tdir = Path(td)
        subprocess.run(
            ["unzip", "-q", "-o", str(args.zip), "-d", str(tdir)],
            check=True,
        )
        html = tdir / "pcb-BOM.html"
        if not html.exists():
            raise SystemExit(f"No pcb-BOM.html in {args.zip}")
        ibom = load_ibom_rows(html)
        cart = load_cart(args.cart)

    lines: list[tuple[int, int, str, str, str, str, str, str, str]] = []
    for idx, (qty, value, fp, refs, extras) in enumerate(ibom, start=1):
        designators = ",".join(refs)
        cart_hit = resolve_cart_keys(extras, cart)
        if cart_hit:
            manufacturer = cart_hit["manufacturer"] or ""
            mpn = cart_hit["mfr_part"] or ""
            desc = cart_hit["description"] or value
        else:
            manufacturer = ""
            mpn = ""
            desc = value
        typ = guess_footprint_type(fp)
        notes_parts: list[str] = []
        if extras:
            notes_parts.append(f"Mouser / schematic: {extras[0]}")
        if not cart_hit and extras and str(extras[0]).strip():
            notes_parts.append("No match in saved cart — fill Manufacturer / Mfg Part # manually.")
        if not extras or not str(extras[0]).strip():
            notes_parts.append("No distributor # in KiCad BOM — match module manually (e.g. ESP32 devkit).")
        notes: str = " ".join(notes_parts) if notes_parts else ""

        lines.append((idx, qty, designators, manufacturer, mpn, desc, fp, typ, notes))

    title = "ESP32 OpenLCB IO Board v1.01 1PCS BOM  (Bill of Materials)"
    write_pcbway_xlsx(args.out, title, lines)
    print(f"Wrote {args.out} ({len(lines)} line items)")
    _maybe_cleanup_venv()


if __name__ == "__main__":
    main()
